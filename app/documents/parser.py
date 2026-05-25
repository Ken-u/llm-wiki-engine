"""Document parsers: PDF, DOCX, XLSX, TXT/MD."""

from __future__ import annotations

from pathlib import Path


def parse_pdf(path: Path) -> str:
    import fitz  # pymupdf

    doc = fitz.open(str(path))
    pages = []
    for page in doc:
        pages.append(page.get_text("text"))
    doc.close()
    return "\n\n".join(pages).strip()


def parse_docx(path: Path) -> str:
    from docx import Document

    doc = Document(str(path))
    return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())


def parse_xlsx(path: Path) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    lines: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"## Sheet: {sheet_name}\n")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            lines.append(" | ".join(cells))
    wb.close()
    return "\n".join(lines)


def parse_text(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace")


_PARSERS = {
    ".pdf": parse_pdf,
    ".docx": parse_docx,
    ".xlsx": parse_xlsx,
    ".xls": parse_xlsx,
    ".txt": parse_text,
    ".md": parse_text,
    ".markdown": parse_text,
    ".rst": parse_text,
    ".csv": parse_text,
    ".json": parse_text,
    ".yaml": parse_text,
    ".yml": parse_text,
}


def parse_document(path: Path) -> str:
    ext = path.suffix.lower()
    parser = _PARSERS.get(ext)
    if parser is None:
        return parse_text(path)
    return parser(path)


def supported_extensions() -> list[str]:
    return list(_PARSERS.keys())
